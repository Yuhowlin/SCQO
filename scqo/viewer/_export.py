"""Setup-page export builders — pure data -> bytes, renderer-free (no fastapi).

Both builders consume the setup page's row dicts
``{entity, field, value, unit, source, previous}`` (:func:`app._param_rows`)
and ignore ``previous`` — an export carries current state only. The xlsx
builder lazy-imports openpyxl (the viewer extra); its ImportError propagates so
the route can answer with the reinstall hint. The PDF builder lazy-imports
headless matplotlib — guaranteed transitively via scqat — per the
``cli/_campaign_plot`` discipline, and renders a 16:9 DOCUMENT twin of the
offline HTML export (same sections, same order, page-style source text),
flowing across pages — not per-source-kind slides.
"""

from __future__ import annotations

import base64
import io

#: one naming authority for the xlsx route's response header
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: the two setup-page sections, in page order — the xlsx sheet names
SECTIONS = ("Calibration", "Physical parameters")

_HEADER = ("entity", "parameter", "value", "unit", "source")

# the viewer palette (base.html / _style.html)
_MUTED, _EDGE, _HEAD = "#6b7683", "#e2e8ef", "#f2f6fa"


def _fmt(v) -> str:
    """The page's value formatting (setup.html): %.8g for numerics, str otherwise."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return f"{v:.8g}"
    return str(v)


def source_label(src: dict | None) -> str:
    """One short filterable string per provenance status (the xlsx source cell)."""
    if not src:
        return "unrecorded"
    status = src.get("status")
    if status == "run":
        return f"run:{src.get('run_id')}"
    if status == "campaign":
        return f"campaign:{src.get('campaign_id')}"
    return status or "unrecorded"


def source_text(src: dict | None) -> str:
    """The page's source column as plain text — the pdf mirrors the html."""
    if not src:
        return "-"
    status = src.get("status")
    if status == "run":
        return src.get("run_id") or "-"
    if status == "campaign":
        return f"{src.get('campaign_id')} (campaign)"
    if status == "manual":
        return "(manual)"
    if status == "external":
        return "(externally changed)"
    return "-"


def xlsx_bytes(state_rows: list[dict], physical_rows: list[dict]) -> bytes:
    """Two-sheet workbook mirroring the page sections, 5 columns each."""
    import openpyxl  # lazy: only the xlsx route pays for it (and can 503 without it)
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    for title, rows in zip(SECTIONS, (state_rows, physical_rows)):
        ws = wb.active if title == SECTIONS[0] else wb.create_sheet()
        ws.title = title
        ws.append(_HEADER)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            v = r["value"]
            if not isinstance(v, (int, float, str, bool)) and v is not None:
                v = str(v)  # list-shaped values exist; keep numerics native
            ws.append((r["entity"], r["field"], v, r["unit"] or "",
                       source_label(r["source"])))
        for col, width in zip("ABCDE", (16, 28, 18, 10, 44)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------- pdf

# 16:9 page geometry (fractions of the page unless noted)
_PAGE = (13.333, 7.5)                    # inches
_X0, _X1 = 0.04, 0.96
_TOP, _BOT = 0.93, 0.075
_RH = 0.038                              # one table row
_GAP = 0.018                             # breathing room after tables/figures
_H_TITLE, _H_H3, _H_H4, _H_CAP = 0.11, 0.055, 0.042, 0.034
_PARAM_WIDTHS = (0.13, 0.24, 0.17, 0.08, 0.38)


def _decode_data_uri(uri: str):
    """data:image/png;base64,... -> image array, or None on anything
    undecodable (a torn or placeholder png must never fail the export)."""
    import matplotlib.image as mpimg
    try:
        raw = base64.b64decode(uri.split(",", 1)[1])
        return mpimg.imread(io.BytesIO(raw), format="png")
    except (ValueError, OSError, IndexError):
        return None


def _param_block(rows: list[dict]) -> dict:
    return {"k": "table", "header": _HEADER, "widths": _PARAM_WIDTHS, "fs": 10,
            "rows": [(r["entity"], r["field"], _fmt(r["value"]),
                      r["unit"] or "", source_text(r["source"])) for r in rows]}


def _stat_cells(s: dict) -> tuple:
    if s.get("nonscalar"):
        return (s["experiment"], s["target"], s["quantity"],
                "-", "-", "(array-valued)", "-", "-", "-", "-", "-")

    def g(key, spec="%.5g"):
        return spec % s[key] if s.get(key) is not None else "-"

    return (s["experiment"], s["target"], s["quantity"],
            str(s.get("n", "-")), str(s.get("n_missing", "-")),
            g("mean"), g("std"), g("sem"), g("min"), g("max"),
            g("scatter_ratio", "%.2f"))


def _pdf_blocks(ctx: dict, embedded_runs: list[dict],
                embedded_campaigns: list[dict], exported_at: str) -> list[dict]:
    """The html export's sections as layout blocks, in the same order."""
    dev, cd, sname = ctx["device"], ctx["cooldown"], ctx["setup_name"]
    active = " · active" if ctx.get("is_active") else ""
    blocks: list[dict] = [
        {"k": "title", "text": f"Setup: {sname}  ({dev} · {cd}{active})",
         "sub": f"Exported {exported_at} — offline copy of the setup page"}]

    meta: list[tuple[str, str]] = []
    cycle = ctx.get("cycle") or {}
    if cycle:
        meta += [("cycle", f"{cd}{' (active)' if ctx.get('is_active') else ''}"),
                 ("start", str(cycle.get("start", ""))),
                 ("end", str(cycle.get("end") or "-"))]
        if cycle.get("fridge"):
            meta.append(("fridge", str(cycle["fridge"])))
        if cycle.get("packaging"):
            meta.append(("packaging", str(cycle["packaging"])))
    sm = ctx.get("setup_meta")
    if sm:
        meta += [("backend", str(sm.get("backend", ""))),
                 ("instrument config", str(sm.get("instrument_config") or "(built-in)"))]
        if sm.get("note"):
            meta.append(("note", str(sm["note"])))
    if meta:
        blocks.append({"k": "table", "header": None, "rows": meta,
                       "widths": (0.22, 0.78), "fs": 10})

    blocks.append({"k": "h3", "text": "Calibration"})
    if ctx.get("authority") == "state":
        blocks.append({"k": "cap", "text":
                       f"Current calibration — from {ctx.get('state_path')}"})
    elif ctx.get("authority") == "snapshot":
        rid = (ctx.get("snapshot_run") or {}).get("run_id", "")
        blocks.append({"k": "cap", "text": f"Last observed calibration — from "
                       f"{rid} (device_after snapshot; no state file yet)"})
    else:
        blocks.append({"k": "cap",
                       "text": "No calibration recorded for this context yet."})
    if ctx["state_rows"]:
        blocks.append(_param_block(ctx["state_rows"]))

    blocks.append({"k": "h3", "text": "Physical parameters"})
    if ctx["physical_rows"]:
        blocks.append(_param_block(ctx["physical_rows"]))
    else:
        blocks.append({"k": "cap", "text":
                       "No physical parameters recorded for this context yet."})
    blocks.append({"k": "cap", "text":
                   "source = the run the CURRENT value came from (strict match); "
                   "(externally changed) = the current value matches no SCQO record."})

    if embedded_runs:
        blocks.append({"k": "h3", "text": "Referenced runs"})
    for r in embedded_runs:
        if not r.get("found"):
            blocks.append({"k": "h4", "text": r["run_id"]})
            blocks.append({"k": "cap", "text":
                           "not found in the index — the run data may have moved."})
            continue
        rec = r["record"]
        badges = ", ".join(f"{q}: {o}"
                           for q, o in (rec.get("outcomes") or {}).items())
        blocks.append({"k": "h4",
                       "text": r["run_id"] + (f"   [{badges}]" if badges else "")})
        line = (f"{rec.get('experiment')} · {rec.get('started_at')} · "
                f"backend {rec.get('backend')}")
        if rec.get("operator"):
            line += f" · operator {rec['operator']}"
        blocks.append({"k": "cap", "text": line})
        fit_rows = [(q, k, _fmt(v)) for q, d in (r.get("fit") or {}).items()
                    for k, v in d.items()]
        if fit_rows:
            blocks.append({"k": "table", "header": ("target", "quantity", "value"),
                           "rows": fit_rows, "widths": (0.3, 0.4, 0.3), "fs": 10})
        for uri in r.get("figures") or ():
            arr = _decode_data_uri(uri)
            if arr is not None:
                blocks.append({"k": "img", "arr": arr})

    if embedded_campaigns:
        blocks.append({"k": "h3", "text": "Referenced campaigns"})
    for c in embedded_campaigns:
        if not c.get("found"):
            blocks.append({"k": "h4", "text": c["campaign_id"]})
            blocks.append({"k": "cap", "text": "not found in the index — "
                           "the campaign data may have moved."})
            continue
        status = (c.get("manifest") or {}).get("status", "")
        blocks.append({"k": "h4", "text":
                       c["campaign_id"] + (f"   ({status})" if status else "")})
        if c.get("stats_rows"):
            blocks.append({"k": "table",
                           "header": ("experiment", "target", "quantity", "n",
                                      "miss", "mean", "std", "sem", "min", "max",
                                      "std/err"),
                           "rows": [_stat_cells(s) for s in c["stats_rows"]],
                           "widths": (0.16, 0.07, 0.13, 0.05, 0.05, 0.10, 0.10,
                                      0.10, 0.09, 0.09, 0.06), "fs": 8})
        if c.get("png"):
            arr = _decode_data_uri(c["png"])
            if arr is not None:
                blocks.append({"k": "img", "arr": arr})
    return blocks


def _paginate(blocks: list[dict]) -> list[list[dict]]:
    """Flow blocks down 16:9 pages: tables split with a repeated header (min 3
    rows before a break), figures never split, headings keep a lead of their
    content. Every placed item carries its top-y."""
    pages: list[list[dict]] = [[]]
    y = _TOP

    def newpage():
        nonlocal y
        pages.append([])
        y = _TOP

    def need(h: float, keep: float = 0.0):
        if y - (h + keep) < _BOT and pages[-1]:
            newpage()

    for b in blocks:
        k = b["k"]
        if k == "title":
            pages[-1].append({**b, "y": y})
            y -= _H_TITLE
        elif k in ("h3", "h4"):
            h = _H_H3 if k == "h3" else _H_H4
            # never strand a heading at a page's foot: a section heading must
            # pull its caption AND a visible bite of its table along
            need(h, keep=0.18 if k == "h3" else 0.12)
            pages[-1].append({**b, "y": y})
            y -= h
        elif k == "cap":
            need(_H_CAP)
            pages[-1].append({**b, "y": y})
            y -= _H_CAP
        elif k == "table":
            rows = list(b["rows"])
            hdr = 1 if b["header"] else 0
            while rows:
                avail = int((y - _BOT) / _RH) - hdr
                if avail < min(3, len(rows)) and pages[-1]:
                    newpage()
                    continue
                avail = max(avail, 1)
                chunk, rows = rows[:avail], rows[avail:]
                n = len(chunk) + hdr
                pages[-1].append({**b, "rows": chunk, "y": y, "n": n})
                y -= n * _RH + _GAP
        elif k == "img":
            ih, iw = b["arr"].shape[0], b["arr"].shape[1]
            wfrac = 0.45
            h = wfrac * (ih / iw) * (_PAGE[0] / _PAGE[1])
            if h > 0.62:  # keep pixel aspect: shrink width with the height cap
                wfrac *= 0.62 / h
                h = 0.62
            need(h)
            pages[-1].append({**b, "y": y, "h": h, "w": wfrac})
            y -= h + _GAP
    return pages


def pdf_bytes(ctx: dict, embedded_runs: list[dict],
              embedded_campaigns: list[dict], *, exported_at: str) -> bytes:
    """The offline HTML export as a 16:9 PDF document: title + meta, the two
    parameter tables (page-style source text), then the referenced runs (fit
    tables + figures) and campaigns. ``ctx`` is _setup_context's dict; the
    embedded lists are the html route's own loader output (figures arrive as
    data: URIs and are decoded back here — one loader, two renderers)."""
    import matplotlib

    matplotlib.use("Agg")  # headless — the _campaign_plot discipline
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    context = f"{ctx['device']} · {ctx['cooldown']}/{ctx['setup_name']}"
    pages = _paginate(_pdf_blocks(ctx, embedded_runs, embedded_campaigns,
                                  exported_at))
    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        for i, items in enumerate(pages):
            fig = plt.figure(figsize=_PAGE)
            for it in items:
                k = it["k"]
                if k == "title":
                    fig.text(_X0, it["y"] - 0.045, it["text"],
                             fontsize=20, fontweight="bold")
                    fig.text(_X0, it["y"] - 0.085, it["sub"],
                             fontsize=10, color=_MUTED)
                elif k == "h3":
                    fig.text(_X0, it["y"] - 0.042, it["text"],
                             fontsize=15, fontweight="bold")
                elif k == "h4":
                    fig.text(_X0, it["y"] - 0.032, it["text"],
                             fontsize=11.5, fontweight="bold")
                elif k == "cap":
                    fig.text(_X0, it["y"] - 0.026, it["text"],
                             fontsize=9, color=_MUTED)
                elif k == "table":
                    h = it["n"] * _RH
                    ax = fig.add_axes([_X0, it["y"] - h, _X1 - _X0, h])
                    ax.axis("off")
                    kw = dict(cellText=[[str(c) for c in row]
                                        for row in it["rows"]],
                              colWidths=list(it["widths"]), cellLoc="left",
                              bbox=[0.0, 0.0, 1.0, 1.0])
                    if it["header"]:
                        kw["colLabels"] = it["header"]
                    tbl = ax.table(**kw)
                    tbl.auto_set_font_size(False)
                    tbl.set_fontsize(it["fs"])
                    for (row, _col), cell in tbl.get_celld().items():
                        cell.set_edgecolor(_EDGE)
                        cell.set_linewidth(0.6)
                        cell.PAD = 0.02
                        if it["header"] and row == 0:
                            cell.set_facecolor(_HEAD)
                            cell.set_text_props(fontweight="bold")
                elif k == "img":
                    ax = fig.add_axes([_X0, it["y"] - it["h"],
                                       it["w"], it["h"]])
                    ax.imshow(it["arr"], aspect="auto")
                    ax.axis("off")
            fig.text(_X0, 0.03, f"SCQO setup export — {context}",
                     fontsize=9, color=_MUTED)
            fig.text(_X1, 0.03, f"{i + 1} / {len(pages)}",
                     ha="right", fontsize=9, color=_MUTED)
            pdf.savefig(fig)
            plt.close(fig)
    return buf.getvalue()
