"""The 3-sheet characterization workbook — pure data -> bytes, no fastapi.

Same contract as :mod:`scqo.viewer._export`: it takes the setup page's context
and returns the file. openpyxl is lazy-imported (the viewer extra) and its
ImportError propagates so the route can answer with the reinstall hint.

Sheet names and the field-dictionary content come from
:mod:`~scqo.viewer.lab_report.template`; the numbers come from
:mod:`~scqo.viewer.lab_report.metrics`. Nothing lab-specific is written here.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from .metrics import extract_chip_metrics
from .template import load_template

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: Data-sheet rows, in sheet order: (label, metrics key or None for a manual row).
#: The Dashboard's cell formulas below are indexed against THIS list, so a row
#: added or moved here moves them too — hence _data_row_index rather than the
#: hardcoded row numbers the first draft carried.
DATA_ROWS: tuple[tuple[str, str | None], ...] = (
    ("Basic info", None),
    ("QPU name", None),
    ("meas person", None),
    ("holder", None),
    ("cooldown_date", None),
    ("DR", None),
    ("OM", None),
    ("Raw data URI", None),
    ("Resonator", "__resonator_name__"),
    ("Bare resonant frequency (GHz)", "f_bare_ghz"),
    ("Q_c", "q_c"),
    ("Q_i_e0", "q_i"),
    ("Q_i_e4", None),
    ("qubit (sweet spot)", "__qubit_name__"),
    ("Junction Resistance (ohm)", None),
    ("Resistance to fq (GHz)", "r_to_fq"),
    ("Dressed resonant frequency (|0>)(GHz)", "f_dress_ghz"),
    ("Qubit frequency (GHz)", "f_q_ghz"),
    ("Tunable", "__tunable__"),
    ("T1 #100 (us)", "t1_mean_us"),
    ("T1_error (us)", "t1_std_us"),
    ("T2_Ramsey #1 (us)", "t2_ramsey_single_us"),
    ("T2_echo #1 (us)", "t2_echo_single_us"),
    ("Readout fidelity", "readout_fidelity_single"),
    ("T2_Ramsey #100 (us)", "t2_ramsey_mean_us"),
    ("T2_Ramsey_error (us)", "t2_ramsey_std_us"),
    ("T2_echo #100 (us)", "t2_echo_mean_us"),
    ("T2_echo_error (us)", "t2_echo_std_us"),
    ("Readout fidelity #100", "readout_fidelity_mean"),
    ("thermal_population #100", "thermal_population_mean"),
    ("RB fidelity #100", "rb_fidelity_mean"),
    ("flux_bias (V)", None),
    ("flux_period (V)", None),
    ("f02/2 (GHz)", "f02_half_ghz"),
    ("Resonator shift Δf_r(|1⟩−|0⟩)", "delta_fr_mhz"),
    ("Resonator linewidth κ/2π", "kappa_mhz"),
    ("Qubit-resonator coupling g/2π", "g_mhz"),
    ("Anharmonicity α/2π", "anharmonicity_mhz"),
    ("EJ/EC", "ej_ec"),
    ("Coupler", None),
    ("Coupler sweetspot frequency", None),
    ("Coupler idle frequency", None),
    ("Residual ZZ/2π", None),
)

#: the Data rows the Dashboard's section headers bold.
_HEADER_LABELS = ("Basic info", "Resonator", "qubit (sweet spot)")


def _data_row_index(label: str) -> int:
    """The 1-based sheet row a Data label lands on. Derived, never written by
    hand: the Dashboard formulas break silently if the two disagree."""
    for i, (name, _key) in enumerate(DATA_ROWS, start=1):
        if name == label:
            return i
    raise KeyError(f"no Data row labelled {label!r}")


def _format_mean_std(mean: float | None, std: float | None, digits: int = 2,
                     is_pct: bool = False) -> str:
    """``mean ± std`` for display, or ``""`` when there is no mean."""
    if mean is None:
        return ""
    factor, suffix = (100.0, "%") if is_pct else (1.0, "")
    out = f"{mean * factor:.{digits}f}"
    if std is not None and std > 0:
        out += f" ± {std * factor:.{digits}f}"
    return out + suffix


def _cell(per_q: dict, qubit: str, key: str | None) -> Any:
    """One Data-sheet cell. Numbers stay NUMBERS so the workbook can compute on
    them; only the special columns render as text."""
    if key is None:
        return ""
    if key == "__resonator_name__" or key == "__qubit_name__":
        return qubit.upper()
    if key == "__tunable__":
        return "TRUE" if per_q[qubit]["tunable"] else "FALSE"
    value = per_q[qubit].get(key)
    return "" if value is None else value


def lab_template_xlsx_bytes(ctx: dict, store: Any = None, data_root: Path | None = None,
                            min_repeats: int = 2, estimator: str = "mean",
                            tags: Any = None, design: Any = None,
                            predicted_f_q: dict[str, float] | None = None) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    tpl = load_template()
    metrics = extract_chip_metrics(
        ctx, store=store, data_root=data_root, min_repeats=min_repeats,
        estimator=estimator, tags=tags, design=design, predicted_f_q=predicted_f_q,
    )
    qubits, per_q = metrics["qubits"], metrics["per_qubit"]
    cycle = metrics["cycle"]

    thin = Side(style="thin", color="DCE4EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", start_color="F2F6FA", end_color="F2F6FA")
    summary_fill = PatternFill("solid", start_color="E9ECEF", end_color="E9ECEF")
    bold, title_font = Font(bold=True), Font(bold=True, size=13)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ Data
    ws = wb.active
    ws.title = tpl.sheet("data", "Data")
    # Rows with no metrics key are the operator's to fill in; the few the
    # cooldown registry already knows are pre-filled in column B.
    manual = {
        "QPU name": metrics["device"],
        "holder": cycle.get("packaging", ""),
        "cooldown_date": str(cycle.get("start", "") or ""),
        "DR": cycle.get("fridge", ""),
    }
    for label, key in DATA_ROWS:
        if key is None:
            value = manual.get(label, "")
            ws.append([label, value] if value else [label])
        else:
            ws.append([label, *[_cell(per_q, q, key) for q in qubits]])
    for label in _HEADER_LABELS:
        row = _data_row_index(label)
        for col in range(1, len(qubits) + 2):
            c = ws.cell(row=row, column=col)
            c.font, c.fill = bold, head_fill

    ws.column_dimensions["A"].width = 38
    for i in range(len(qubits)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    # ------------------------------------------------------------- Dashboard
    dash = wb.create_sheet(title=tpl.sheet("dashboard", "Dashboard"))
    dash["A1"] = "QPU Characterization Dashboard"
    dash["A1"].font = title_font

    headers = ["Qubit", "Resistance to fq (GHz)", "Qubit frequency (GHz)", "Resonator",
               "Design resonator frequency (GHz)", "Bare resonant frequency (GHz)",
               "Predicted high", "Predicted low", "Measured high", "Measured low",
               "Design high", "Design low", "Bare high", "Bare low",
               "Qubit plot domain", "Resonator plot domain"]
    dash.append(headers)
    for col in range(1, len(headers) + 1):
        c = dash.cell(row=2, column=col)
        c.font, c.fill = bold, head_fill

    data_sheet = ws.title
    r_qubit = _data_row_index("qubit (sweet spot)")
    r_pred = _data_row_index("Resistance to fq (GHz)")
    r_fq = _data_row_index("Qubit frequency (GHz)")
    r_res = _data_row_index("Resonator")
    r_bare = _data_row_index("Bare resonant frequency (GHz)")

    def pick(a: str, b: str, row: int, take: str, cmp: str) -> str:
        return (f'=IF(AND(ISNUMBER({a}{row}),ISNUMBER({b}{row})),'
                f'IF({a}{row}{cmp}{b}{row},{take}{row},""),"")')

    for i, q in enumerate(qubits):
        row = i + 3
        col = get_column_letter(i + 2)
        design_bare = per_q[q].get("design_f_bare_ghz")
        dash.append([
            f"='{data_sheet}'!{col}{r_qubit}",
            f"='{data_sheet}'!{col}{r_pred}",
            f"='{data_sheet}'!{col}{r_fq}",
            f"='{data_sheet}'!{col}{r_res}",
            design_bare if design_bare is not None else "",
            f"='{data_sheet}'!{col}{r_bare}",
            pick("B", "C", row, "B", ">"), pick("B", "C", row, "B", "<="),
            pick("C", "B", row, "C", ">"), pick("C", "B", row, "C", "<="),
            pick("E", "F", row, "E", ">="), pick("E", "F", row, "E", "<"),
            pick("F", "E", row, "F", ">"), pick("F", "E", row, "F", "<="),
            f"=A{row}", f"=D{row}",
        ])

    for _ in range(max(0, 29 - dash.max_row)):
        dash.append([])

    dash.append(["Characterization summary", *[q.upper() for q in qubits]])
    for col in range(1, len(qubits) + 2):
        c = dash.cell(row=dash.max_row, column=col)
        c.font, c.fill, c.border = bold, summary_fill, border

    def num(v, digits=2):
        return f"{v:.{digits}f}" if isinstance(v, (int, float)) else ("" if v is None else str(v))

    def prefer(mean_key, single_key, label_100, label_1, digits=2, pct=False):
        """The #100 statistic when any qubit has one, else the single shot."""
        has = any(per_q[q][mean_key] is not None for q in qubits)
        vals = [
            _format_mean_std(per_q[q][mean_key], per_q[q][mean_key.replace("_mean_", "_std_")],
                             digits, pct)
            if per_q[q][mean_key] is not None
            else (_format_mean_std(per_q[q][single_key], None, digits, pct) if pct
                  else num(per_q[q][single_key], digits))
            for q in qubits
        ]
        return (label_100 if has else label_1), vals

    rows: list[tuple[str, list[str]]] = [
        ("Dressed resonant frequency (|0>) (GHz)", [num(per_q[q]["f_dress_ghz"], 3) for q in qubits]),
        ("Qubit frequency (GHz)", [num(per_q[q]["f_q_ghz"], 3) for q in qubits]),
        ("T1 #100 (us)", [_format_mean_std(per_q[q]["t1_mean_us"], per_q[q]["t1_std_us"]) for q in qubits]),
        ("T2*/T1", [num(per_q[q]["t2_star_over_t1"]) for q in qubits]),
        ("T2e/T1", [num(per_q[q]["t2_echo_over_t1"]) for q in qubits]),
        ("Tunable", ["TRUE" if per_q[q]["tunable"] else "FALSE" for q in qubits]),
        prefer("t2_ramsey_mean_us", "t2_ramsey_single_us", "T2_Ramsey #100 (us)", "T2_Ramsey #1 (us)"),
        prefer("t2_echo_mean_us", "t2_echo_single_us", "T2_echo #100 (us)", "T2_echo #1 (us)"),
        prefer("readout_fidelity_mean", "readout_fidelity_single",
               "Readout fidelity #100", "Readout fidelity", pct=True),
        ("Temperature #100 (mk)",
         [_format_mean_std(per_q[q]["temperature_mk"], per_q[q]["temperature_err_mk"], 1) for q in qubits]),
        ("thermal_population #100",
         [_format_mean_std(per_q[q]["thermal_population_mean"], per_q[q]["thermal_population_std"], 2, True)
          for q in qubits]),
        ("RB fidelity #100",
         [_format_mean_std(per_q[q]["rb_fidelity_mean"], per_q[q]["rb_fidelity_std"], 2, True) for q in qubits]),
        ("Resonator shift Δfr (MHz)", [num(per_q[q]["delta_fr_mhz"], 3) for q in qubits]),
        ("κ/2π (MHz)", [num(per_q[q]["kappa_mhz"], 3) for q in qubits]),
        ("g/2π (MHz, approx.)", [num(per_q[q]["g_mhz"], 3) for q in qubits]),
        ("α/2π (MHz)", [num(per_q[q]["anharmonicity_mhz"]) for q in qubits]),
        ("EJ/EC", [num(per_q[q]["ej_ec"]) for q in qubits]),
    ]
    for label, values in rows:
        dash.append([label, *values])
        r = dash.max_row
        for col in range(1, len(values) + 2):
            c = dash.cell(row=r, column=col)
            c.border = border
            if col == 1:
                c.font = bold
            else:
                c.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------ Field dictionary
    dic = wb.create_sheet(title=tpl.sheet("dictionary", "Field dictionary"))
    dic.append([tpl.sheet("dictionary_title", "")])
    dic["A1"].font = title_font
    dic.append(list(tpl.dictionary_header))
    for col in range(1, len(tpl.dictionary_header) + 1):
        c = dic.cell(row=2, column=col)
        c.font, c.fill = bold, head_fill
    for row in tpl.dictionary:
        dic.append(row.as_cells())
    for letter, width in zip("ABCDEFG", (15, 8, 32, 12, 14, 65, 15)):
        dic.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
