"""Lab-standard 3-sheet Excel export builder (Data, Dashboard, Field Dictionary).

Generates a clean, robust openpyxl workbook matching the lab's characterization template:
1. 'Data': Raw parameters per qubit/resonator, basic metadata, and derived physics quantities.
2. 'Dashboard': QPU Characterization Dashboard with comparison columns and
   the Characterization Summary table linking cleanly to 'Data'.
3. 'Field Dictionary': Standard dictionary defining each parameter and field.
"""

from __future__ import annotations

import csv
import io
import math
from pathlib import Path
from typing import Any

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Field dictionary definitions (matching standard lab dictionary)
FIELD_DICTIONARY_ROWS = [
    ("基本資訊", 2.0, "QPU name", "手動", "文字", "QPU／sample 唯一名稱；建議與資料夾命名一致。", "P0｜白"),
    ("基本資訊", 3.0, "meas person", "手動", "文字", "主要量測人員；多人可用逗號分隔。", "P0｜白"),
    ("基本資訊", 4.0, "holder", "手動", "文字", "Sample holder／封裝座編號。", "P0｜白"),
    ("基本資訊", 5.0, "cooldown_date", "手動", "日期", "建議 YYYY-MM-DD；既有 YYMMDD 格式可沿用但需一致。", "P0｜白"),
    ("基本資訊", 6.0, "DR", "手動", "文字", "Dilution refrigerator／量測冰箱識別。", "P0｜白"),
    ("基本資訊", 7.0, "OM", "手動", "URL", "Optical microscope／製程檢查資料連結。", "P0｜白"),
    ("基本資訊", 8.0, "Raw data URI", "手動", "URI", "原始量測資料位置；建議可解析且長期不變。", "P0｜白"),
    ("Resonator", 9.0, "Resonator", "結構", "元件 ID", "B:Z 的 resonator 欄名；衍生式要求與 row 14 qubit 同欄配對。", "P0｜白"),
    ("Resonator", 10.0, "Bare resonant frequency", "手動", "GHz", "未耦合／基準 resonator 頻率；需註明取得方法與 dressed 值同條件。", "P0｜白"),
    ("Resonator", 11.0, "Q_c", "手動", "無因次", "Coupling quality factor。", "P3｜深灰"),
    ("Resonator", 12.0, "Q_i_e0", "手動", "無因次", "Internal Q at e0 regime；請在 SOP 固定 e0 的 photon number／drive-power 定義。", "P3｜深灰"),
    ("Resonator", 13.0, "Q_i_e4", "手動", "無因次", "Internal Q at e4 regime；請在 SOP 固定 e4 的 photon number／drive-power 定義。", "P3｜深灰"),
    ("Qubit", 14.0, "qubit (sweet spot)", "結構", "元件 ID", "B:Z 的 qubit 欄名；固定頻率 qubit 亦以工作點填入。", "P0｜白"),
    ("Qubit", 15.0, "Junction Resistance", "手動", "ohm", "Josephson junction normal resistance；外部來源若為 kΩ 必須 ×1000。", "P0｜白"),
    ("Qubit", 16.0, "Resistance to fq", "手動", "GHz", "由 junction resistance／模型得到的預估 f01；模型版本與 Ec 假設需留在來源註記。", "P0｜白"),
    ("Qubit", 17.0, "Dressed resonant frequency (|0>)", "手動", "GHz", "含 qubit 耦合後的 readout resonator 頻率。", "P0｜白"),
    ("Qubit", 18.0, "Qubit frequency", "手動", "GHz", "量測 f01；對 tunable qubit 預設記 sweet spot，其他 bias 點另留 raw data。", "P0｜白"),
    ("Qubit", 19.0, "Tunable", "手動", "TRUE/FALSE", "是否可藉 flux 調頻。", "P0｜白"),
    ("Qubit", 20.0, "T1 #100", "手動", "us", "100 次重複量測的 T1 統計代表值；需在 SOP 定義 mean／median。", "P0｜白"),
    ("Qubit", 21.0, "T1_error", "手動", "us", "T1 #100 的誤差；需在 SOP 明確定義 SD、SEM 或 fit uncertainty。", "P0｜白"),
    ("Qubit", 22.0, "T2_Ramsey #1", "手動", "us", "單次 Ramsey dephasing time（常記 T2*）。", "P0｜白"),
    ("Qubit", 23.0, "T2_echo #1", "手動", "us", "單次 Hahn-echo coherence time。", "P0｜白"),
    ("Qubit", 24.0, "Readout fidelity", "手動", "0–1", "單次 readout assignment fidelity；需固定 estimator 與 calibration protocol。", "P0｜白"),
    ("Qubit", 25.0, "T2_Ramsey #100", "手動", "us", "100 次 Ramsey 的統計代表值；需在 SOP 定義 mean／median。", "P1｜淺灰"),
    ("Qubit", 26.0, "T2_Ramsey_error", "手動", "us", "T2_Ramsey #100 的誤差；需定義 SD／SEM／fit uncertainty。", "P1｜淺灰"),
    ("Qubit", 27.0, "T2_echo #100", "手動", "us", "100 次 echo 的統計代表值。", "P1｜淺灰"),
    ("Qubit", 28.0, "T2_echo_error", "手動", "us", "T2_echo #100 的誤差；需定義 SD／SEM／fit uncertainty。", "P1｜淺灰"),
    ("Qubit", 29.0, "Readout fidelity #100", "手動", "0–1", "100 次重複校正／量測的 readout fidelity；#100 protocol 需固定。", "P1｜淺灰"),
    ("Qubit", 30.0, "thermal_population #100", "手動", "0–1", "100 次估計的 excited-state thermal population。", "P1｜淺灰"),
    ("Qubit", 31.0, "RB fidelity #100", "手動", "0–1", "100 次 randomized benchmarking 的 gate fidelity；需記 Clifford／primitive 定義。", "P1｜淺灰"),
    ("Qubit", 32.0, "flux_bias", "手動", "V", "選定工作點的 flux-bias voltage。", "P2｜中灰"),
    ("Qubit", 33.0, "flux_period", "手動", "V", "一個 flux quantum 對應的 bias period。", "P2｜中灰"),
    ("Qubit", 34.0, "f02/2", "手動", "GHz", "二光子 0→2 躍遷頻率的一半；用於 α 與 EJ/EC 衍生。", "P2｜中灰"),
    ("Qubit", 35.0, "Resonator shift Δf_r(|1⟩−|0⟩)", "被動", "MHz", "2χ=2g²α/[Δ(Δ+α)]，Δ=fq−fr,bare；有號值、dispersive 近似。", "衍生｜黃"),
    ("Qubit", 36.0, "Resonator linewidth κ/2π", "被動", "MHz", "1000×fr,dressed×(1/Qc+1/Qi)；Qi 優先 e4，否則 e0。", "衍生｜黃"),
    ("Qubit", 37.0, "Qubit-resonator coupling g/2π", "被動", "MHz", "1000×√| (fr,dressed−fr,bare)(fr,dressed−fq) |；二能級模態估算。", "衍生｜黃"),
    ("Qubit", 38.0, "Anharmonicity α/2π", "被動", "MHz", "2000×(f02/2−f01)。", "衍生｜黃"),
    ("Qubit", 39.0, "EJ/EC", "被動", "無因次", "[(1000f01/(−α)+1)²]/8；transmon 近似且 α<0。", "衍生｜黃"),
    ("Coupler", 40.0, "Coupler", "結構", "元件 ID", "B:Z 的 coupler／edge 欄名；建議以 Q1–Q2 等 pair ID 命名。", "P0｜白"),
    ("Coupler", 41.0, "Coupler sweetspot frequency", "手動", "GHz", "Coupler sweet-spot 頻率。", "P0｜白"),
    ("Coupler", 42.0, "Coupler idle frequency", "手動", "GHz", "Coupler idle operating frequency。", "P0｜白"),
    ("Coupler", 43.0, "Residual ZZ/2π", "手動", "kHz", "Coupler／qubit pair 在 idle point 的 residual ZZ。", "P1｜淺灰"),
]


try:
    import tomllib
except ModuleNotFoundError:
    import tomli as tomllib  # type: ignore[no-redef]


def _read_design_bare_freq(path: Path) -> dict[str, float]:
    """Read bare resonator design frequencies in GHz from design.toml."""
    if not path.is_file():
        return {}
    try:
        with open(path, "rb") as fp:
            data = tomllib.load(fp)
        res: dict[str, float] = {}
        for entity_name, fields in data.items():
            if isinstance(fields, dict):
                hz_val = fields.get("f_bare_hz") or fields.get("f_bare")
                if isinstance(hz_val, (int, float)):
                    base = entity_name.split("_")[0].lower()
                    res[base] = hz_val / 1e9
        return res
    except Exception:
        return {}


def _read_r_to_fq(csv_path: Path) -> dict[str, float]:
    """Read R_to_fq.csv file mapping qubit names to predicted fq in GHz."""
    if not csv_path.is_file():
        return {}
    res: dict[str, float] = {}
    try:
        with csv_path.open("r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 2:
                    k, v = row[0].strip().lower(), row[1].strip()
                    try:
                        res[k] = float(v)
                    except ValueError:
                        pass
    except Exception:
        pass
    return res


from scqo.campaign_query import query_campaign_statistics


def extract_chip_metrics(ctx: dict, store: Any = None, data_root: Path | None = None,
                         min_repeats: int = 2, estimator: str = "mean",
                         tags: Any = None) -> dict[str, Any]:
    """Extract metrics across all qubits in the context, combining physical parameters,
    state knobs/monitors, and independent campaign statistics."""
    dev = ctx.get("device", "")
    cd = ctx.get("cooldown", "")
    sname = ctx.get("setup_name", "")
    cycle = ctx.get("cycle") or {}

    state_rows = ctx.get("state_rows") or []
    phys_rows = ctx.get("physical_rows") or []

    # Map (entity, field) -> value
    state_map: dict[tuple[str, str], Any] = {(r["entity"], r["field"]): r["value"] for r in state_rows}
    phys_map: dict[tuple[str, str], Any] = {(r["entity"], r["field"]): r["value"] for r in phys_rows}

    # Discover qubit names across all entities in this context
    qubits: list[str] = []
    for r in phys_rows + state_rows:
        ent = r["entity"]
        base = ent.split("_")[0]
        if base.startswith("q") and base not in qubits:
            qubits.append(base)

    def qkey(x: str):
        num = "".join(c for c in x if c.isdigit())
        return (0 if num else 1, int(num) if num else x)

    qubits = sorted(qubits, key=qkey) if qubits else ["q1"]

    # Check for optional R_to_fq.csv prediction table
    r_to_fq: dict[str, float] = {}
    design_bare: dict[str, float] = {}
    if data_root is not None:
        cand_csv = data_root / dev / "R_to_fq.csv"
        if not cand_csv.is_file():
            cand_csv = data_root / dev / cd / sname / "R_to_fq.csv"
        r_to_fq = _read_r_to_fq(cand_csv)

        cand_design = data_root / dev / "design.toml"
        if not cand_design.is_file():
            cand_design = data_root / dev / cd / sname / "design.toml"
        design_bare = _read_design_bare_freq(cand_design)

    # Query latest qualifying campaign statistics using the high-level Campaign Query Engine
    campaign_stats = query_campaign_statistics(
        store,
        device=dev,
        cooldown=cd if cd else None,
        min_repeats=min_repeats,
        status=("complete", "running"),
        estimator=estimator,  # type: ignore[arg-type]
        tags=tags,
    )
    if not campaign_stats:
        campaign_stats = query_campaign_statistics(
            store,
            device=dev,
            min_repeats=min_repeats,
            status=("complete", "running"),
            estimator=estimator,  # type: ignore[arg-type]
            tags=tags,
        )

    per_qubit: dict[str, dict[str, Any]] = {}
    for q in qubits:
        q_lower = q.lower()
        q_res = f"{q}_res"
        q_ro = f"{q}_ro"
        q_xy = f"{q}_xy"
        q_z = f"{q}_z"

        # 1. Resonator Frequencies (GHz)
        f_dress = phys_map.get((q_res, "f_dress0_hz")) or state_map.get((q_ro, "readout_freq_hz"))
        f_dress_ghz = (f_dress / 1e9) if isinstance(f_dress, (int, float)) else None

        f_bare = phys_map.get((q_res, "f_bare_hz"))
        f_bare_ghz = (f_bare / 1e9) if isinstance(f_bare, (int, float)) else None

        # 2. Qubit Frequency (GHz)
        f_q = phys_map.get((q, "f_01_hz")) or phys_map.get((q, "f_q_max_hz")) or state_map.get((q_xy, "drive_freq_hz"))
        f_q_ghz = (f_q / 1e9) if isinstance(f_q, (int, float)) else None

        # 3. Tunable status
        tunable = True if (q_z, "idle_flux") in state_map or phys_map.get((q, "f_q_max_hz")) is not None else False

        cq = campaign_stats.get(q_lower, {})

        # 4. T1 (us) - Single measurement vs Campaign statistics
        t1_val = phys_map.get((q, "t1_s"))
        t1_single_us = (t1_val * 1e6) if isinstance(t1_val, (int, float)) else None

        t1_camp = cq.get("t1_s") or cq.get("t1")
        if t1_camp is not None:
            t1_mean_us = t1_camp["value"] * 1e6
            t1_std_us = (t1_camp["error"] * 1e6) if t1_camp.get("error") is not None else None
            t1_n = t1_camp.get("n")
        else:
            t1_mean_us = None
            t1_std_us = None
            t1_n = None

        # 5. T2_Ramsey (us) - Single measurement vs Campaign statistics
        t2_ramsey_val = phys_map.get((q, "t2_star_s"))
        t2_ramsey_single_us = (t2_ramsey_val * 1e6) if isinstance(t2_ramsey_val, (int, float)) else None

        t2_ram_camp = cq.get("t2_star_s") or cq.get("t2_star")
        if t2_ram_camp is not None:
            t2_ramsey_mean_us = t2_ram_camp["value"] * 1e6
            t2_ramsey_std_us = (t2_ram_camp["error"] * 1e6) if t2_ram_camp.get("error") is not None else None
            t2_ramsey_n = t2_ram_camp.get("n")
        else:
            t2_ramsey_mean_us = None
            t2_ramsey_std_us = None
            t2_ramsey_n = None

        # 6. T2_echo (us) - Single measurement vs Campaign statistics
        t2_echo_val = phys_map.get((q, "t2_echo_s"))
        t2_echo_single_us = (t2_echo_val * 1e6) if isinstance(t2_echo_val, (int, float)) else None

        t2_echo_camp = cq.get("t2_echo_s") or cq.get("t2_echo")
        if t2_echo_camp is not None:
            t2_echo_mean_us = t2_echo_camp["value"] * 1e6
            t2_echo_std_us = (t2_echo_camp["error"] * 1e6) if t2_echo_camp.get("error") is not None else None
            t2_echo_n = t2_echo_camp.get("n")
        else:
            t2_echo_mean_us = None
            t2_echo_std_us = None
            t2_echo_n = None

        # Coherence ratios: use priority (#100 if available, else #1)
        eff_t1 = t1_mean_us if t1_mean_us is not None else t1_single_us
        eff_t2_ramsey = t2_ramsey_mean_us if t2_ramsey_mean_us is not None else t2_ramsey_single_us
        eff_t2_echo = t2_echo_mean_us if t2_echo_mean_us is not None else t2_echo_single_us

        t2_star_over_t1 = (eff_t2_ramsey / eff_t1) if (eff_t2_ramsey and eff_t1 and eff_t1 > 0) else None
        t2_echo_over_t1 = (eff_t2_echo / eff_t1) if (eff_t2_echo and eff_t1 and eff_t1 > 0) else None

        # 7. Readout Fidelity (Single-shot state discrimination assignment fidelity)
        fid_g = state_map.get((q_ro, "fidelity_g"))
        fid_e = state_map.get((q_ro, "fidelity_e"))
        if isinstance(fid_g, (int, float)) and isinstance(fid_e, (int, float)):
            ro_fid_single = (fid_g + fid_e) / 2.0
        elif isinstance(fid_g, (int, float)):
            ro_fid_single = fid_g
        elif isinstance(fid_e, (int, float)):
            ro_fid_single = fid_e
        else:
            ro_fid_single = state_map.get((q_ro, "readout_fidelity"))

        ro_camp = cq.get("readout_fidelity") or cq.get("fidelity_g")
        if ro_camp is not None:
            ro_fid_mean = ro_camp["value"]
            ro_fid_std = ro_camp.get("error")
            ro_fid_n = ro_camp.get("n")
        else:
            ro_fid_mean = None
            ro_fid_std = None
            ro_fid_n = None

        # 8. Thermal Population (n_th from physical.json mode fact)
        nth_single = phys_map.get((q, "n_th"))
        nth_camp = cq.get("n_th") or cq.get("pop_e_prep_g")
        if nth_camp is not None:
            nth_mean = nth_camp.get("value") or nth_camp.get("mean")
            nth_std = nth_camp.get("error") or nth_camp.get("std")
        else:
            nth_mean = nth_single
            nth_std = None

        # 9. Effective Temperature (mK) derived from n_th and f_01
        temp_mk = None
        temp_err_mk = None
        if nth_mean is not None and isinstance(nth_mean, (int, float)) and nth_mean > 0 and f_q_ghz:
            try:
                # T_eff = h * f01 / (k_B * ln(1 + 1/n_th)) * 1000 mK
                denom = math.log(1.0 + (1.0 / nth_mean))
                temp_mk = (47.9924 * f_q_ghz) / denom
                if nth_std is not None and nth_std > 0:
                    temp_err_mk = temp_mk * (nth_std / (nth_mean * (1.0 + nth_mean) * denom))
            except (ValueError, ZeroDivisionError):
                pass

        # 10. Randomized Benchmarking Fidelity
        rb_single = phys_map.get((q, "rb_fidelity")) or state_map.get((q, "rb_fidelity"))
        rb_camp = cq.get("fidelity") or cq.get("rb_fidelity")
        if rb_camp is not None:
            rb_fid_mean = rb_camp.get("value") or rb_camp.get("mean")
            rb_fid_std = rb_camp.get("error") or rb_camp.get("std")
        else:
            rb_fid_mean = rb_single
            rb_fid_std = None

        # 11. Anharmonicity (MHz) & Quality factors
        anharm = phys_map.get((q, "anharmonicity_hz"))
        anharm_mhz = (anharm / 1e6) if isinstance(anharm, (int, float)) else None

        q_c = phys_map.get((q_res, "q_c"))
        q_i = phys_map.get((q_res, "q_i"))

        # 12. Derived physical quantities: g/2pi, kappa/2pi, delta_fr, ej_ec
        g_mhz = None
        if f_dress_ghz is not None and f_bare_ghz is not None and f_q_ghz is not None:
            diff = abs((f_dress_ghz - f_bare_ghz) * (f_dress_ghz - f_q_ghz))
            g_mhz = 1000.0 * math.sqrt(diff)

        kappa_mhz = None
        if f_dress_ghz is not None:
            inv_qc = (1.0 / q_c) if (isinstance(q_c, (int, float)) and q_c > 0) else 0.0
            inv_qi = (1.0 / q_i) if (isinstance(q_i, (int, float)) and q_i > 0) else 0.0
            if inv_qc > 0 or inv_qi > 0:
                kappa_mhz = 1000.0 * f_dress_ghz * (inv_qc + inv_qi)

        delta_fr_mhz = None
        if f_q_ghz is not None and f_bare_ghz is not None and g_mhz is not None and anharm_mhz:
            delta_q = (f_q_ghz - f_bare_ghz) * 1000.0
            denom = delta_q * (delta_q + anharm_mhz)
            if denom != 0:
                delta_fr_mhz = (2.0 * (g_mhz ** 2) * anharm_mhz) / denom

        ej_ec = None
        if f_q_ghz is not None and anharm_mhz and anharm_mhz < 0:
            ej_ec = (((1000.0 * f_q_ghz / (-anharm_mhz)) + 1.0) ** 2) / 8.0

        f02_half_ghz = ((f_q_ghz * 1000.0 + (anharm_mhz / 2.0)) / 1000.0) if (f_q_ghz is not None and anharm_mhz is not None) else None

        pred_fq = r_to_fq.get(q.lower()) or r_to_fq.get(q.upper())
        des_f_bare = design_bare.get(q.lower()) or design_bare.get(q.upper())

        per_qubit[q] = {
            "design_f_bare_ghz": des_f_bare,
            "f_dress_ghz": f_dress_ghz,
            "f_bare_ghz": f_bare_ghz,
            "f_q_ghz": f_q_ghz,
            "f02_half_ghz": f02_half_ghz,
            "tunable": tunable,
            "t1_single_us": t1_single_us,
            "t1_mean_us": t1_mean_us,
            "t1_std_us": t1_std_us,
            "t1_n": t1_n,
            "t2_ramsey_single_us": t2_ramsey_single_us,
            "t2_ramsey_mean_us": t2_ramsey_mean_us,
            "t2_ramsey_std_us": t2_ramsey_std_us,
            "t2_ramsey_n": t2_ramsey_n,
            "t2_echo_single_us": t2_echo_single_us,
            "t2_echo_mean_us": t2_echo_mean_us,
            "t2_echo_std_us": t2_echo_std_us,
            "t2_echo_n": t2_echo_n,
            "t2_star_over_t1": t2_star_over_t1,
            "t2_echo_over_t1": t2_echo_over_t1,
            "readout_fidelity_single": ro_fid_single,
            "readout_fidelity_mean": ro_fid_mean,
            "readout_fidelity_std": ro_fid_std,
            "thermal_population_mean": nth_mean,
            "thermal_population_std": nth_std,
            "temperature_mk": temp_mk,
            "temperature_err_mk": temp_err_mk,
            "rb_fidelity_single": rb_single,
            "rb_fidelity_mean": rb_fid_mean,
            "rb_fidelity_std": rb_fid_std,
            "anharmonicity_mhz": anharm_mhz,
            "q_c": q_c,
            "q_i": q_i,
            "g_mhz": g_mhz,
            "kappa_mhz": kappa_mhz,
            "delta_fr_mhz": delta_fr_mhz,
            "ej_ec": ej_ec,
            "r_to_fq": pred_fq,
        }

    return {
        "device": dev,
        "cooldown": cd,
        "setup_name": sname,
        "cycle": cycle,
        "qubits": qubits,
        "per_qubit": per_qubit,
    }


def _format_mean_std(mean: float | None, std: float | None, digits: int = 2, is_pct: bool = False) -> str:
    """Format mean ± std string for dashboard display."""
    if mean is None:
        return ""
    factor = 100.0 if is_pct else 1.0
    suffix = "%" if is_pct else ""
    m_val = mean * factor
    if std is not None and std > 0:
        s_val = std * factor
        return f"{m_val:.{digits}f} ± {s_val:.{digits}f}{suffix}"
    return f"{m_val:.{digits}f}{suffix}"


def lab_template_xlsx_bytes(ctx: dict, store: Any = None, data_root: Path | None = None,
                            min_repeats: int = 2, estimator: str = "mean",
                            tags: Any = None) -> bytes:
    """Generate a 3-sheet Excel workbook matching the lab's characterization template."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    metrics = extract_chip_metrics(
        ctx, store=store, data_root=data_root,
        min_repeats=min_repeats, estimator=estimator, tags=tags,
    )
    qubits = metrics["qubits"]
    per_q = metrics["per_qubit"]
    dev = metrics["device"]
    cycle = metrics["cycle"]
    cooldown_date = cycle.get("start", "")
    fridge = cycle.get("fridge", "")
    holder = cycle.get("packaging", "")

    wb = openpyxl.Workbook()

    thin_border = Border(
        left=Side(style="thin", color="DCE4EC"),
        right=Side(style="thin", color="DCE4EC"),
        top=Side(style="thin", color="DCE4EC"),
        bottom=Side(style="thin", color="DCE4EC"),
    )
    header_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    summary_hdr_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=13)

    # ------------------------------------------------------------- 1. Data Sheet
    ws_data = wb.active
    ws_data.title = "Data"

    data_rows: list[list[Any]] = [
        ["Basic info"],
        ["QPU name", dev],
        ["meas person", ""],
        ["holder", holder],
        ["cooldown_date", str(cooldown_date) if cooldown_date else ""],
        ["DR", fridge],
        ["OM", ""],
        ["Raw data URI", str(data_root / dev if data_root else dev)],
        ["Resonator", *[q.upper() for q in qubits]],
        ["Bare resonant frequency (GHz)", *[per_q[q]["f_bare_ghz"] for q in qubits]],
        ["Q_c", *[per_q[q]["q_c"] for q in qubits]],
        ["Q_i_e0", *[per_q[q]["q_i"] for q in qubits]],
        ["Q_i_e4", *["" for _ in qubits]],
        ["qubit (sweet spot)", *[q.upper() for q in qubits]],
        ["Junction Resistance (ohm)", *["" for _ in qubits]],
        ["Resistance to fq (GHz)", *[per_q[q]["r_to_fq"] for q in qubits]],
        ["Dressed resonant frequency (|0>)(GHz)", *[per_q[q]["f_dress_ghz"] for q in qubits]],
        ["Qubit frequency (GHz)", *[per_q[q]["f_q_ghz"] for q in qubits]],
        ["Tunable", *["TRUE" if per_q[q]["tunable"] else "FALSE" for q in qubits]],
        ["T1 #100 (us)", *[per_q[q]["t1_mean_us"] for q in qubits]],
        ["T1_error (us)", *[per_q[q]["t1_std_us"] for q in qubits]],
        ["T2_Ramsey #1 (us)", *[per_q[q]["t2_ramsey_single_us"] for q in qubits]],
        ["T2_echo #1 (us)", *[per_q[q]["t2_echo_single_us"] for q in qubits]],
        ["Readout fidelity", *[per_q[q]["readout_fidelity_single"] for q in qubits]],
        ["T2_Ramsey #100 (us)", *[per_q[q]["t2_ramsey_mean_us"] for q in qubits]],
        ["T2_Ramsey_error (us)", *[per_q[q]["t2_ramsey_std_us"] for q in qubits]],
        ["T2_echo #100 (us)", *[per_q[q]["t2_echo_mean_us"] for q in qubits]],
        ["T2_echo_error (us)", *[per_q[q]["t2_echo_std_us"] for q in qubits]],
        ["Readout fidelity #100", *[per_q[q]["readout_fidelity_mean"] for q in qubits]],
        ["thermal_population #100", *[per_q[q]["thermal_population_mean"] for q in qubits]],
        ["RB fidelity #100", *[per_q[q]["rb_fidelity_mean"] for q in qubits]],
        ["flux_bias (V)", *["" for _ in qubits]],
        ["flux_period (V)", *["" for _ in qubits]],
        ["f02/2 (GHz)", *[per_q[q]["f02_half_ghz"] for q in qubits]],
        ["Resonator shift Δf_r(|1⟩−|0⟩)", *[per_q[q]["delta_fr_mhz"] for q in qubits]],
        ["Resonator linewidth κ/2π", *[per_q[q]["kappa_mhz"] for q in qubits]],
        ["Qubit-resonator coupling g/2π", *[per_q[q]["g_mhz"] for q in qubits]],
        ["Anharmonicity α/2π", *[per_q[q]["anharmonicity_mhz"] for q in qubits]],
        ["EJ/EC", *[per_q[q]["ej_ec"] for q in qubits]],
        ["Coupler", *["" for _ in qubits]],
        ["Coupler sweetspot frequency", *["" for _ in qubits]],
        ["Coupler idle frequency", *["" for _ in qubits]],
        ["Residual ZZ/2π", *["" for _ in qubits]],
    ]

    for r_idx, row in enumerate(data_rows, start=1):
        ws_data.append(row)
        if r_idx in (1, 9, 14):
            for c_idx in range(1, len(row) + 1):
                cell = ws_data.cell(row=r_idx, column=c_idx)
                cell.font = bold_font
                cell.fill = header_fill

    num_q = len(qubits)
    ws_data.column_dimensions["A"].width = 38
    for i in range(num_q):
        ws_data.column_dimensions[get_column_letter(i + 2)].width = 16

    # -------------------------------------------------------- 2. Dashboard Sheet
    ws_dash = wb.create_sheet(title="Dashboard")

    ws_dash["A1"] = "QPU Characterization Dashboard"
    ws_dash["A1"].font = title_font

    top_headers = [
        "Qubit", "Resistance to fq (GHz)", "Qubit frequency (GHz)", "Resonator",
        "Design resonator frequency (GHz)", "Bare resonant frequency (GHz)",
        "Predicted high", "Predicted low", "Measured high", "Measured low",
        "Design high", "Design low", "Bare high", "Bare low",
        "Qubit plot domain", "Resonator plot domain"
    ]
    ws_dash.append(top_headers)
    for c_idx in range(1, len(top_headers) + 1):
        cell = ws_dash.cell(row=2, column=c_idx)
        cell.font = bold_font
        cell.fill = header_fill

    for i, q in enumerate(qubits):
        r_num = i + 3
        col_letter = get_column_letter(i + 2)
        des_bare_val = per_q[q].get("design_f_bare_ghz")
        row_cells = [
            f"=Data!{col_letter}14",
            f"=Data!{col_letter}16",
            f"=Data!{col_letter}18",
            f"=Data!{col_letter}9",
            des_bare_val if des_bare_val is not None else "",
            f"=Data!{col_letter}10",
            f'=IF(AND(ISNUMBER(B{r_num}),ISNUMBER(C{r_num})),IF(B{r_num}>C{r_num},B{r_num},""),"")',
            f'=IF(AND(ISNUMBER(B{r_num}),ISNUMBER(C{r_num})),IF(B{r_num}<=C{r_num},B{r_num},""),"")',
            f'=IF(AND(ISNUMBER(B{r_num}),ISNUMBER(C{r_num})),IF(C{r_num}>B{r_num},C{r_num},""),"")',
            f'=IF(AND(ISNUMBER(B{r_num}),ISNUMBER(C{r_num})),IF(C{r_num}<=B{r_num},C{r_num},""),"")',
            f'=IF(AND(ISNUMBER(E{r_num}),ISNUMBER(F{r_num})),IF(E{r_num}>=F{r_num},E{r_num},""),"")',
            f'=IF(AND(ISNUMBER(E{r_num}),ISNUMBER(F{r_num})),IF(E{r_num}<F{r_num},E{r_num},""),"")',
            f'=IF(AND(ISNUMBER(E{r_num}),ISNUMBER(F{r_num})),IF(F{r_num}>E{r_num},F{r_num},""),"")',
            f'=IF(AND(ISNUMBER(E{r_num}),ISNUMBER(F{r_num})),IF(F{r_num}<=E{r_num},F{r_num},""),"")',
            f"=A{r_num}",
            f"=D{r_num}",
        ]
        ws_dash.append(row_cells)

    for _ in range(max(0, 29 - ws_dash.max_row)):
        ws_dash.append([])

    summary_header = ["Characterization summary", *[q.upper() for q in qubits]]
    ws_dash.append(summary_header)
    hdr_row_idx = ws_dash.max_row
    for c_idx in range(1, len(summary_header) + 1):
        cell = ws_dash.cell(row=hdr_row_idx, column=c_idx)
        cell.font = bold_font
        cell.fill = summary_hdr_fill
        cell.border = thin_border

    def fmt_num(v, digits=2):
        return f"{v:.{digits}f}" if isinstance(v, (int, float)) else ("" if v is None else str(v))

    has_ramsey_100 = any(per_q[q]["t2_ramsey_mean_us"] is not None for q in qubits)
    ramsey_title = "T2_Ramsey #100 (us)" if has_ramsey_100 else "T2_Ramsey #1 (us)"
    ramsey_vals = [
        _format_mean_std(per_q[q]["t2_ramsey_mean_us"], per_q[q]["t2_ramsey_std_us"], 2)
        if per_q[q]["t2_ramsey_mean_us"] is not None
        else fmt_num(per_q[q]["t2_ramsey_single_us"], 2)
        for q in qubits
    ]

    has_echo_100 = any(per_q[q]["t2_echo_mean_us"] is not None for q in qubits)
    echo_title = "T2_echo #100 (us)" if has_echo_100 else "T2_echo #1 (us)"
    echo_vals = [
        _format_mean_std(per_q[q]["t2_echo_mean_us"], per_q[q]["t2_echo_std_us"], 2)
        if per_q[q]["t2_echo_mean_us"] is not None
        else fmt_num(per_q[q]["t2_echo_single_us"], 2)
        for q in qubits
    ]

    has_ro_100 = any(per_q[q]["readout_fidelity_mean"] is not None for q in qubits)
    ro_title = "Readout fidelity #100" if has_ro_100 else "Readout fidelity"
    ro_vals = [
        _format_mean_std(per_q[q]["readout_fidelity_mean"], per_q[q]["readout_fidelity_std"], 2, is_pct=True)
        if per_q[q]["readout_fidelity_mean"] is not None
        else _format_mean_std(per_q[q]["readout_fidelity_single"], None, 2, is_pct=True)
        for q in qubits
    ]
    summary_rows_def = [
        ("Dressed resonant frequency (|0>) (GHz)",
         [fmt_num(per_q[q]["f_dress_ghz"], 3) for q in qubits]),
        ("Qubit frequency (GHz)",
         [fmt_num(per_q[q]["f_q_ghz"], 3) for q in qubits]),
        ("T1 #100 (us)",
         [_format_mean_std(per_q[q]["t1_mean_us"], per_q[q]["t1_std_us"], 2) for q in qubits]),
        ("T2*/T1",
         [fmt_num(per_q[q]["t2_star_over_t1"], 2) for q in qubits]),
        ("T2e/T1",
         [fmt_num(per_q[q]["t2_echo_over_t1"], 2) for q in qubits]),
        ("Tunable",
         ["TRUE" if per_q[q]["tunable"] else "FALSE" for q in qubits]),
        (ramsey_title, ramsey_vals),
        (echo_title, echo_vals),
        (ro_title, ro_vals),
        ("Temperature #100 (mk)",
         [_format_mean_std(per_q[q]["temperature_mk"], per_q[q]["temperature_err_mk"], 1) for q in qubits]),
        ("thermal_population #100",
         [_format_mean_std(per_q[q]["thermal_population_mean"], per_q[q]["thermal_population_std"], 2, is_pct=True) for q in qubits]),
        ("RB fidelity #100",
         [_format_mean_std(per_q[q]["rb_fidelity_mean"], per_q[q]["rb_fidelity_std"], 2, is_pct=True) for q in qubits]),
        ("Resonator shift Δfr (MHz)",
         [fmt_num(per_q[q]["delta_fr_mhz"], 3) for q in qubits]),
        ("κ/2π (MHz)",
         [fmt_num(per_q[q]["kappa_mhz"], 3) for q in qubits]),
        ("g/2π (MHz, approx.)",
         [fmt_num(per_q[q]["g_mhz"], 3) for q in qubits]),
        ("α/2π (MHz)",
         [fmt_num(per_q[q]["anharmonicity_mhz"], 2) for q in qubits]),
        ("EJ/EC",
         [fmt_num(per_q[q]["ej_ec"], 2) for q in qubits]),
    ]

    for title, row_data in summary_rows_def:
        row_vals = [title, *row_data]
        ws_dash.append(row_vals)
        cur_row = ws_dash.max_row
        for c_idx in range(1, len(row_vals) + 1):
            cell = ws_dash.cell(row=cur_row, column=c_idx)
            cell.border = thin_border
            if c_idx == 1:
                cell.font = bold_font
            else:
                cell.alignment = Alignment(horizontal="center")


    # ------------------------------------------------- 3. Field Dictionary Sheet
    ws_dict = wb.create_sheet(title="欄位字典")
    ws_dict.append(["QPU 公版欄位字典（標準規格）"])
    ws_dict["A1"].font = title_font
    ws_dict.append(["區段", "項次", "欄位名稱", "輸入方式", "型態/單位", "說明", "分類/重要度"])

    for c_idx in range(1, 8):
        cell = ws_dict.cell(row=2, column=c_idx)
        cell.font = bold_font
        cell.fill = header_fill

    for r_data in FIELD_DICTIONARY_ROWS:
        ws_dict.append(list(r_data))

    ws_dict.column_dimensions["A"].width = 15
    ws_dict.column_dimensions["B"].width = 8
    ws_dict.column_dimensions["C"].width = 32
    ws_dict.column_dimensions["D"].width = 12
    ws_dict.column_dimensions["E"].width = 14
    ws_dict.column_dimensions["F"].width = 65
    ws_dict.column_dimensions["G"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
